#!/usr/bin/env python3
"""Merge verified-*.json device batches into device-compatibility.xlsx."""
import glob
import json
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

WORKDIR = "/Users/adnaan/Documents/hl-homelander-app/.devicelist-work"
OUT = "/Users/adnaan/Documents/hl-homelander-app/device-compatibility.xlsx"

TIER_ORDER = {"Supported": 0, "Likely": 1, "Unverified": 2}
TIER_FILL = {
    "Supported": PatternFill("solid", fgColor="C6EFCE"),
    "Likely": PatternFill("solid", fgColor="FFEB9C"),
    "Unverified": PatternFill("solid", fgColor="D9D9D9"),
}
TIER_FONT = {
    "Supported": Font(color="276221", bold=True),
    "Likely": Font(color="9C6500", bold=True),
    "Unverified": Font(color="3F3F3F", bold=True),
}
GATE_COLOR = {"pass": "276221", "likely": "9C6500", "unknown": "808080", "fail": "9C0006"}

COLS = [
    ("Tier", "tier", 12),
    ("Make", "make", 12),
    ("Model", "model", 26),
    ("Model numbers (IN/BR)", "model_numbers", 28),
    ("Codename", "codename", 13),
    ("Year", "launch_year", 7),
    ("SoC", "soc", 26),
    ("Android (launch→max)", "_android", 12),
    ("Ultrawide camera", "ultrawide_sensor", 24),
    ("UW FOV (°)", "ultrawide_fov_deg", 9),
    ("Gate: UW ≥110° dFOV", "gate_ultrawide_dfov", 11),
    ("Gate: HEVC 1080p30 CBR", "gate_hevc_1080p30_cbr", 11),
    ("Gate: IMU ≥100 Hz", "gate_imu_100hz", 11),
    ("Gate: REALTIME ts", "gate_realtime_timestamp", 11),
    ("Gate: GMS/Integrity", "gate_gms_play_integrity", 11),
    ("Gate: 25-min thermal", "gate_thermal_25min", 11),
    ("India", "_in", 7),
    ("Brazil", "_br", 7),
    ("~Price (₹)", "price_inr", 16),
    ("~Price (R$)", "price_brl", 16),
    ("Caveats", "caveats", 60),
    ("Notes", "notes", 50),
    ("Sources", "_sources", 45),
]

ABOUT = [
    ("Humyn Labs Capture — Android device compatibility list", True),
    ("Generated 2026-06-12 from a multi-source research sweep (GSMArena/Kimovil/OEM pages, ARCore certification list, "
     "developer-community Camera2 probe evidence, SoC encoder reports) + in-house hardware testing.", False),
    ("", False),
    ("SCOPE: Android slab phones (no foldables/tablets) launched Jan 2020 – Jun 2026, officially sold in India or Brazil, "
     "any price. Devices without a rear ultrawide, without GMS, or stuck on Android 10 or lower are excluded outright.", False),
    ("", False),
    ("TIERS", True),
    ("Supported — high confidence on ALL gates: in-house tested, or flagship/upper-midrange with documented ultrawide "
     "≥118°, known-good SoC encoder, ARCore-certified.", False),
    ("Likely — passes on paper with 1–2 gates unverifiable from public data (e.g. ultrawide exposure to third-party "
     "apps unconfirmed, marginal 112–117° FOV, mid-tier MediaTek encoder).", False),
    ("Unverified — meets paper spec but multiple significant unknowns (budget IMU, possible UNKNOWN camera timestamp "
     "source, encoder risk). Verify on hardware before publishing as supported.", False),
    ("", False),
    ("HARD GATES (the app's runtime compat-check; verdicts per device: pass / likely / unknown / fail)", True),
    ("1. UW ≥110° dFOV — rear ultrawide computing to ≥110° diagonal FOV, reachable by third-party Camera2 apps via "
     "CONTROL_ZOOM_RATIO (requires Android 11+).", False),
    ("2. HEVC 1080p30 CBR — hardware HEVC encode at exactly 1920×1080/30fps honoring CBR with zero B-frames.", False),
    ("3. IMU ≥100 Hz — gyroscope + accelerometer sustaining ≥100 Hz with p99 inter-sample interval ≤12 ms under "
     "camera load.", False),
    ("4. REALTIME ts — Camera2 SENSOR_INFO_TIMESTAMP_SOURCE = REALTIME.", False),
    ("5. GMS/Integrity — Google Mobile Services + Play Integrity (excludes post-2019 Huawei).", False),
    ("6. 25-min thermal — sustain 25 min of 1080p30 HEVC capture without severe throttling (Pixel 7a-class reference).", False),
    ("", False),
    ("Ground truth: Pixel 7a / 8a / 10a verified passing on in-house hardware. Galaxy S22 Exynos variant flagged for "
     "HEVC encoder timestamp instability (Snapdragon variant fine).", False),
    ("A device on this list is a *candidate*: the app still runs its own on-device compat-check at first launch and is "
     "the final authority. Prices are approximate street prices as of June 2026.", False),
]


def norm_key(d):
    return re.sub(r"\s+", " ", f"{d.get('make','')}|{d.get('model','')}".lower()).strip()


def main():
    rows, seen = [], {}
    for f in sorted(glob.glob(f"{WORKDIR}/verified-*.json")):
        try:
            batch = json.load(open(f))
        except Exception as e:
            print(f"WARN: cannot parse {f}: {e}")
            continue
        if not isinstance(batch, list):
            print(f"WARN: {f} is not a list, skipping")
            continue
        for d in batch:
            if not isinstance(d, dict) or not d.get("make") or not d.get("model"):
                continue
            k = norm_key(d)
            if k in seen:  # keep the richer row
                prev = seen[k]
                if sum(1 for v in d.values() if v) > sum(1 for v in prev.values() if v):
                    rows[rows.index(prev)] = d
                    seen[k] = d
                continue
            seen[k] = d
            rows.append(d)

    rows.sort(key=lambda d: (TIER_ORDER.get(d.get("tier"), 3), str(d.get("make", "")).lower(),
                             -(d.get("launch_year") or 0), str(d.get("model", ""))))

    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"
    thin = Border(bottom=Side(style="thin", color="DDDDDD"))
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(color="FFFFFF", bold=True, size=10)

    for c, (title, _, width) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill, cell.font = head_fill, head_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30

    for r, d in enumerate(rows, 2):
        d["_android"] = " → ".join(x for x in [str(d.get("android_launch") or ""), str(d.get("android_max") or "")] if x) or None
        d["_in"] = "Yes" if d.get("available_india") else "No"
        d["_br"] = "Yes" if d.get("available_brazil") else "No"
        src = d.get("sources") or []
        d["_sources"] = "\n".join(src if isinstance(src, list) else [str(src)])
        for c, (_, key, _w) in enumerate(COLS, 1):
            v = d.get(key)
            cell = ws.cell(row=r, column=c, value=v if v not in ("", []) else None)
            cell.border = thin
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("caveats", "notes", "_sources", "model_numbers", "ultrawide_sensor"))
            if key == "tier" and v in TIER_FILL:
                cell.fill, cell.font = TIER_FILL[v], TIER_FONT[v]
            elif key.startswith("gate_") and isinstance(v, str) and v in GATE_COLOR:
                cell.font = Font(color=GATE_COLOR[v], bold=(v == "fail"), size=10)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif key in ("launch_year", "ultrawide_fov_deg", "_in", "_br"):
                cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows) + 1}"

    about = wb.create_sheet("About")
    about.column_dimensions["A"].width = 130
    for i, (text, bold) in enumerate(ABOUT, 1):
        cell = about.cell(row=i, column=1, value=text)
        cell.font = Font(bold=bold, size=12 if i == 1 else 10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT)
    tiers = {}
    for d in rows:
        tiers[d.get("tier", "?")] = tiers.get(d.get("tier", "?"), 0) + 1
    makes = {}
    for d in rows:
        makes[d.get("make", "?")] = makes.get(d.get("make", "?"), 0) + 1
    print(f"Wrote {OUT}: {len(rows)} devices | tiers={tiers}")
    print("by make:", dict(sorted(makes.items(), key=lambda kv: -kv[1])))


if __name__ == "__main__":
    main()

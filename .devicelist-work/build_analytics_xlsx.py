#!/usr/bin/env python3
"""Build the analytics events / funnels workbook for Humyn Labs Capture."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/Users/adnaan/Documents/hl-homelander-app/analytics-events-funnels.xlsx"

# ---- palette ---------------------------------------------------------------
NAVY     = "1F2A44"   # header band
SLATE    = "33415C"   # sub-band
NEW_BG   = "FFF3D6"   # NEW events
EXIST_BG = "E2EFDA"   # existing events
AUTO_BG  = "DDEBF7"   # GA4 auto / derived
SRV_BG   = "F2E2F3"   # server events
ZEBRA    = "F6F8FB"
WHITE    = "FFFFFF"

thin = Side(style="thin", color="D0D7E2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

H_FONT   = Font(name="Calibri", size=11, bold=True, color=WHITE)
T_FONT   = Font(name="Calibri", size=10, color="1A1A1A")
MONO     = Font(name="Consolas", size=10, color="0B3D91")
MONO_B   = Font(name="Consolas", size=10, bold=True, color="0B3D91")
TITLE_F  = Font(name="Calibri", size=16, bold=True, color=NAVY)
SUB_F    = Font(name="Calibri", size=10, italic=True, color="55606E")

TOP  = Alignment(vertical="top", wrap_text=True)
TOPL = Alignment(vertical="top", horizontal="left", wrap_text=True)
CTR  = Alignment(vertical="center", horizontal="center", wrap_text=True)


def style_status(cell, status):
    s = status.lower()
    if "new" in s:
        cell.fill = PatternFill("solid", fgColor=NEW_BG)
    elif "exist" in s:
        cell.fill = PatternFill("solid", fgColor=EXIST_BG)
    elif "auto" in s or "deriv" in s:
        cell.fill = PatternFill("solid", fgColor=AUTO_BG)
    elif "server" in s:
        cell.fill = PatternFill("solid", fgColor=SRV_BG)


def make_table(ws, headers, rows, widths, start_row=1, mono_col=0):
    """Generic table writer. rows = list of tuples. mono_col = index of event-name col."""
    r = start_row
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = H_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = CTR
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[r].height = 26
    r += 1
    for i, row in enumerate(rows):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = TOP
            if c - 1 == mono_col:
                cell.font = MONO
            else:
                cell.font = T_FONT
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ZEBRA)
        # status coloring: assume a column literally named 'Status'
        if "Status" in headers:
            sidx = headers.index("Status")
            style_status(ws.cell(row=r, column=sidx + 1), str(row[sidx]))
        r += 1
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return r


def banner(ws, title, subtitle, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_F
    c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUB_F
    c2.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[2].height = 18


wb = Workbook()

# ===========================================================================
# SHEET 1 — Overview
# ===========================================================================
ws = wb.active
ws.title = "Overview"
ws.sheet_view.showGridLines = False
banner(ws, "Humyn Labs Capture — Analytics Events & Funnels",
       "Spec for GA4 + BigQuery export · setUserId(user_id) + installation_id · client + server for the money path", 2)

rows = [
    ("Destination", "Firebase Analytics (GA4) + BigQuery export. Event/param names kept within GA4 limits (<=40 char names, <=25 params/event). Funnels built in GA4 Funnel Exploration; deep funnels + server joins in BigQuery SQL."),
    ("Identity", "setUserId(<server user UUID, NOT raw Google sub>). User properties: installation_id, app_flavor, consent_version, locale, compat_passed. GA4 auto-collects country / device model / app version."),
    ("Money-path source", "Client logs the full journey; the server independently records upload/finalize/eviction outcomes (Postgres analytics_events table, joined in BigQuery). Client upload_completed will undercount on flaky networks — server is the trustworthy tail."),
    ("Funnels in scope", "1) Activation  2) Task discovery  3) Capture success (money path)  4) Retention / repeat capture."),
    ("Current state of code", "apps/mobile/src/util/analytics.ts already has a FROZEN 47-event allowlist + PII guard, BUT logEvent() is a stub — events only hit the local telemetry ring for /feedback diagnostics; Firebase is never called. This spec = wire the stub to GA4, keep the 47, and fill the gaps (Tasks screen + the entire upload pipeline are dark today)."),
    ("Legend", "Status colors:  NEW = not in code yet (amber) · EXISTS = already in the 47-event allowlist (green) · GA4 AUTO / DERIVED = no client code, from SDK or BigQuery (blue) · SERVER = backend-emitted (purple)."),
    ("PII guard (keep)", "OK: ids (user_id, installation_id, recording_id, task_id), durations, byte sizes, counts, numeric statuses, network_type. NEVER: name, email, task NAME, search QUERY text, recording filename, lat/lng."),
    ("Common params", "Wrapper auto-attaches network_type (wifi|cellular|offline) to every event, and is_practice to every recording/upload event (missing today — practice vs real captures are currently indistinguishable)."),
]
r = 4
for k, v in rows:
    kc = ws.cell(row=r, column=1, value=k)
    kc.font = Font(name="Calibri", size=10, bold=True, color=NAVY)
    kc.alignment = TOPL
    kc.fill = PatternFill("solid", fgColor="EEF1F6")
    kc.border = BORDER
    vc = ws.cell(row=r, column=2, value=v)
    vc.font = T_FONT
    vc.alignment = TOPL
    vc.border = BORDER
    r += 1
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 120
# rough row heights
for rr in range(4, r):
    ws.row_dimensions[rr].height = 46

# ===========================================================================
# Funnel sheets share these headers
# ===========================================================================
FH = ["Step", "Event Name", "Status", "Source", "Attributes (params)", "Funnel role / notes"]
FW = [6, 34, 14, 12, 50, 52]


# --- Funnel 1: Activation ---------------------------------------------------
ws = wb.create_sheet("1 · Activation Funnel")
ws.sheet_view.showGridLines = False
act = [
    ("1",  "first_open",                      "GA4 AUTO", "SDK",    "—", "GA4 install milestone."),
    ("1",  "session_start",                   "GA4 AUTO", "SDK",    "—", "Session boundary (also feeds retention)."),
    ("2",  "splash_shown",                    "EXISTS",   "Client", "—", "App launched."),
    ("3",  "locale_chosen",                   "EXISTS",   "Client", "installation_id, chosen_locale", "First-launch language pick (Phase 7)."),
    ("4",  "signup_started",                  "EXISTS",   "Client", "—", "Signup screen reached."),
    ("4",  "signup_terms_opened",             "EXISTS",   "Client", "—", "Opened the Terms modal."),
    ("4",  "consent_shown",                   "NEW",      "Client", "consent_version", "Consent modal displayed — without this you can't measure scroll-gate dropoff."),
    ("4",  "signup_consent_checked",          "EXISTS",   "Client", "—", "Scroll-gated Agree enabled."),
    ("4",  "consent_agreed",                  "EXISTS",   "Client", "consent_version, time_to_agree_ms (ADD)", "Consent persisted; gates the Continue-with-Google button."),
    ("5",  "signup_google_started",           "EXISTS",   "Client", "—", "Google Sign-In sheet opened."),
    ("5",  "signup_google_completed",         "EXISTS",   "Client", "—", "JWT minted (Play Integrity wrapped inside)."),
    ("5",  "signup_google_failed",            "EXISTS",   "Client", "reason", "Sign-in error / cancel."),
    ("5",  "signup_device_evicted_notice",    "EXISTS",   "Client", "—", "Newest-login-wins: this device superseded an older session (also a retention input)."),
    ("6",  "permission_camera_requested/_granted/_denied", "EXISTS", "Client", "result", "Camera gate."),
    ("6",  "permission_mic_requested/_granted/_denied",    "EXISTS", "Client", "result", "Mic gate."),
    ("6",  "permission_location_requested/_granted/_denied","EXISTS","Client", "result (ADD result=partial)", "Precise-location gate (Bug 3/D4). Coarse-only grant => partial recovery state (Bug 1)."),
    ("6",  "permission_settings_opened",      "NEW",      "Client", "permission", "'Open Settings' recovery path after a denial."),
    ("7",  "compat_started",                  "EXISTS",   "Client", "—", "EncoderProbe -> ImuProbe(30s) -> DeviceCaps begins."),
    ("7",  "compat_check_passed",             "EXISTS",   "Client", "—", "Device meets the locked capture spec."),
    ("7",  "compat_check_failed",             "EXISTS",   "Client", "fail_reason: encoder|imu_hz|device_caps (ADD)", "TERMINAL device rejection — the key India/Brazil fleet-kill signal."),
    ("7",  "compat_completed",               "EXISTS",   "Client", "—", "Compat flow finished."),
    ("8",  "battery_exemption_requested/_granted/_denied", "NEW", "Client", "—", "CompatPass battery-optimisation ask (moved here from Permissions)."),
    ("9",  "rig_tutorial_shown",              "EXISTS",   "Client", "—", "Rig walkthrough."),
    ("9",  "rig_no_rig_link_tapped",          "EXISTS",   "Client", "—", "User has no rig."),
    ("10", "practice_intro_shown",            "EXISTS",   "Client", "—", "Practice intro."),
    ("10", "practice_started",                "EXISTS",   "Client", "—", "60s practice capture begins."),
    ("10", "practice_complete_shown",         "EXISTS",   "Client", "—", "Practice finished."),
    ("10", "practice_complete_continued",     "EXISTS",   "Client", "—", "Reset to MainTabs — onboarding done."),
    ("11", "srv_user_signed_in",              "SERVER",   "Server", "user_id, installation_id, is_new_user, evicted_previous", "Authoritative activation anchor."),
    ("12", "FIRST UPLOAD (milestone)",        "DERIVED",  "BigQuery","first upload_completed / srv_recording_finalized per user_id", "Activation = first successful upload. Derive in BigQuery — the client can't reliably know 'first ever' across reinstalls; the server can."),
]
make_table(ws, FH, act, FW, start_row=1, mono_col=1)

# --- Funnel 2: Task discovery ----------------------------------------------
ws = wb.create_sheet("2 · Task Discovery Funnel")
ws.sheet_view.showGridLines = False
disc = [
    ("1", "tasks_view",                "NEW", "Client", "source: tab|home_tile", "Tasks surface opened."),
    ("2", "task_category_selected",    "NEW", "Client", "category", "Category pill tapped."),
    ("2", "task_list_paginated",       "NEW", "Client", "page, category", "Scrolled to next page (50 of 65)."),
    ("2", "task_search_performed",     "NEW", "Client", "query_length, result_count, latency_ms", "Lexical search (debounced — log on results, NOT per keystroke; no query text per PII guard)."),
    ("2", "task_search_no_results",    "NEW", "Client", "query_length", "Empty search result — catalog-gap signal."),
    ("3", "task_card_tapped",          "NEW", "Client", "task_id, category, position, source: browse|search", "Card tapped in the grid."),
    ("3", "task_details_viewed",       "NEW", "Client", "task_id, category, source", "Details sheet opened."),
    ("3", "task_request_sheet_opened", "NEW", "Client", "source: footer|empty_state", "Opened 'request a task' — demand signal."),
    ("3", "task_request_submitted",    "NEW", "Client", "source", "Submitted a task request."),
    ("4", "task_capture_started",      "NEW", "Client", "task_id, category, source", "CONVERSION: handoff into the Recording screen."),
]
make_table(ws, FH, disc, FW, start_row=1, mono_col=1)

# --- Funnel 3: Capture success ---------------------------------------------
ws = wb.create_sheet("3 · Capture Success Funnel")
ws.sheet_view.showGridLines = False
cap = [
    ("1",  "recording_screen_opened",  "NEW",    "Client", "task_id, is_practice", "Recording modal opened."),
    ("2",  "rotate_prompt_shown",      "NEW",    "Client", "—", "Portrait -> rotate-to-landscape gate."),
    ("2",  "landscape_detected",       "NEW",    "Client", "time_to_landscape_ms", "Rig-fumbling friction before Start is even possible."),
    ("3",  "record_start_pressed",     "NEW",    "Client", "task_id, is_practice", "User tapped record (true 'attempt' start)."),
    ("3",  "pre_flight_failed",        "NEW",    "Client", "reason: thermal|storage|battery", "Device distress kicked back to 'ready'."),
    ("4",  "recording_gate_started",   "EXISTS", "Client", "locale, task_id (ADD), is_practice (ADD)", "MediaPipe hand-gate poll begins."),
    ("4",  "recording_gate_passed",    "EXISTS", "Client", "gate_wait_ms (ADD), poll_count (ADD), miss_count (ADD)", "2 consecutive 2-hand frames confirmed — rig-placement UX quality."),
    ("4",  "recording_gate_skipped",   "EXISTS", "Client", "gate_wait_ms (ADD)", "User tapped Skip (HAND-10) — how long they suffered first."),
    ("4",  "recording_gate_bypassed",  "EXISTS", "Client", "—", "HandDetector native module unavailable."),
    ("5",  "recording_start_failed",   "NEW",    "Client", "reason", "CAPTURE_START_FAILED -> back to 'ready' (invisible today)."),
    ("5",  "recording_started",        "EXISTS", "Client", "recording_id (ADD), task_id (ADD), is_practice (ADD)", "Encoder up, first frame written."),
    ("5",  "recording_orientation_lost","NEW",   "Client", "substate", "Left landscape mid-pre-record; gate reset."),
    ("5",  "battery_alert_shown",      "NEW",    "Client", "elapsed_ms", "Battery overlay during capture."),
    ("5",  "thermal_alert_shown",      "NEW",    "Client", "elapsed_ms", "Thermal overlay during capture."),
    ("6",  "recording_stopped",        "EXISTS", "Client", "reason, duration_ms (ADD), segment_count (ADD), task_id (ADD), is_practice (ADD)", "reason: background|orientation|phone_call|battery_critical|storage_full|permission_revoked|thermal|practice_hard_cap|logout."),
    ("6",  "recording_stop_failed",    "EXISTS", "Client", "reason", "HumynCapture.stop() rejected."),
    ("7",  "segment_finalized",        "NEW",    "Client", "recording_id, segment_index, duration_ms, mean_fps, drift_max_ms, drift_p99_ms", "FinalizeWorker passed. Carries the fleet-health drift telemetry (2026-05-12 owner decision) for free."),
    ("7",  "segment_canceled",         "NEW",    "Client", "recording_id, reason, duration_ms, mean_fps", "reason: fps_dropped|resolution_dropped|insufficient_frames|too_short. The cancel-gate dropout step."),
    ("7",  "recording_too_short",      "EXISTS", "Client", "—", "<3 min non-practice OR <60s practice."),
    ("8",  "upload_enqueued",          "NEW",    "Client", "recording_id, task_id, bytes_total, duration_s", "Row added to on-device queue (practice refused at enqueue, D-08)."),
    ("8",  "upload_started",           "NEW",    "Client", "recording_id, network_type, attempt_count", "Multipart upload begins."),
    ("8",  "upload_retry",             "NEW",    "Client", "recording_id, attempt_count, failure_state, failure_reason", "Auto-retry (e.g. failure_state=FINALIZING)."),
    ("8",  "upload_paused",            "NEW",    "Client", "reason: recording|auth|connectivity", "Queue paused."),
    ("8",  "upload_resumed",           "NEW",    "Client", "reason: recording|auth|connectivity", "Queue resumed."),
    ("8",  "upload_auth_failure",      "NEW",    "Client", "slug: device-evicted|reauth-required|unknown", "401 on init/parts/finalize; row parked, queue paused."),
    ("8",  "upload_needs_attention",   "NEW",    "Client", "recording_id, attempt_count, last_failure_reason", "Auto-retries exhausted; manual retry available."),
    ("8",  "upload_dead_letter",       "NEW",    "Client", "recording_id, dead_letter_reason", "Permanent rejection (409/403/missing bundle)."),
    ("8",  "history_row_retry",        "EXISTS", "Client", "recording_id, reason", "Manual reviveDeadLetter / retryNeedsAttention."),
    ("9",  "upload_completed",         "NEW",    "Client", "recording_id, bytes_total, elapsed_ms, attempt_count, network_type", "TERMINAL SUCCESS on /finalize 200 (client-side)."),
    ("9",  "srv_recording_finalized",  "SERVER", "Server", "user_id, recording_id, bytes, duration_s, ms_since_init", "Trustworthy funnel tail — does not undercount on flaky networks."),
]
make_table(ws, FH, cap, FW, start_row=1, mono_col=1)

# --- Funnel 4: Retention ----------------------------------------------------
ws = wb.create_sheet("4 · Retention Funnel")
ws.sheet_view.showGridLines = False
ret = [
    ("1", "session_start",              "GA4 AUTO", "SDK",     "—", "Return-visit signal -> D1/D7/D30 retention."),
    ("2", "home_view",                  "EXISTS",   "Client",  "—", "Returning user lands on Home."),
    ("2", "home_tile_filter_changed",   "EXISTS",   "Client",  "tile, value", "Engagement with the contribution tiles."),
    ("3", "history_view",               "EXISTS",   "Client",  "—", "Checks on prior contributions."),
    ("3", "history_filter_changed",     "EXISTS",   "Client",  "value", "today|yesterday|this_week|this_month|all|custom."),
    ("3", "history_row_opened",         "EXISTS",   "Client",  "recording_id", "Opened a past segment."),
    ("3", "pending_uploads_view",       "NEW",      "Client",  "pending_count", "Opened the pending-uploads queue screen."),
    ("4", "TIME TO 2ND UPLOAD",         "DERIVED",  "BigQuery","interval between 1st and 2nd srv_recording_finalized", "Core repeat-capture activation metric."),
    ("4", "UPLOADS / USER / WEEK",      "DERIVED",  "BigQuery","count(srv_recording_finalized) by user_id, week", "Contributor productivity / stickiness."),
    ("5", "srv_device_evicted",         "SERVER",   "Server",  "user_id", "Newest-login-wins kicked this user; -> churn if no return session."),
    ("5", "profile_logout",             "EXISTS",   "Client",  "—", "Voluntary churn signal."),
    ("5", "profile_delete_requested",   "EXISTS",   "Client",  "—", "Account-deletion intent."),
    ("5", "profile_delete_confirmed",   "EXISTS",   "Client",  "—", "Hard churn."),
]
make_table(ws, FH, ret, FW, start_row=1, mono_col=1)

# ===========================================================================
# SHEET — Server events (canonical)
# ===========================================================================
ws = wb.create_sheet("Server Events")
ws.sheet_view.showGridLines = False
SH = ["Event Name", "Trigger / endpoint", "Attributes", "Why server-side"]
SW = [26, 34, 50, 60]
srv = [
    ("srv_user_signed_in",      "POST /auth/google", "user_id, installation_id, is_new_user, evicted_previous", "Authoritative activation anchor; flags whether this sign-in evicted a prior device (phone upgrade vs credential sharing)."),
    ("srv_consent_accepted",    "POST /auth/google (consent_log)", "user_id, consent_version", "LEGAL-02 authoritative consent record."),
    ("srv_recording_init",      "POST /recordings/init", "user_id, recording_id, task_id, bytes_expected, has_calibration", "Upload intent that the client may never report if it dies mid-upload."),
    ("srv_recording_finalized", "POST /recordings/{id}/finalize (200)", "user_id, recording_id, bytes, duration_s, ms_since_init", "The trustworthy funnel tail; client upload_completed undercounts on flaky networks/app kills."),
    ("srv_device_evicted",      "requireAuth 401 (installationId mismatch)", "user_id", "Only the server sees the eviction; key churn input."),
    ("srv_feedback_received",   "POST /feedback", "user_id, category", "Confirms feedback landed even if the client request looked like it failed."),
]
make_table(ws, SH, srv, SW, start_row=1, mono_col=0)
# add note row
note_r = ws.max_row + 2
ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=4)
nc = ws.cell(row=note_r, column=1,
             value="Transport recommendation: write to a Postgres analytics_events table "
                   "(event_name, user_id, installation_id, recording_id, props jsonb, created_at) "
                   "and join in BigQuery — rather than GA4 Measurement Protocol, which lacks geo/device "
                   "context and complicates GA4-UI funnels.")
nc.font = SUB_F
nc.alignment = TOPL
ws.row_dimensions[note_r].height = 56

# ===========================================================================
# SHEET — Conventions & Identity
# ===========================================================================
ws = wb.create_sheet("Conventions & Identity")
ws.sheet_view.showGridLines = False
banner(ws, "Conventions, Identity & User Properties", "Apply to every event above.", 2)
conv = [
    ("setUserId", "Server user UUID (NOT the raw Google sub). Set on sign-in, cleared on logout."),
    ("User property: installation_id", "UUID generated at install (MMKV INSTALLATION_ID). Lets you see reinstall / device-eviction churn."),
    ("User property: app_flavor", "apk | playstore — needed to exclude internal/test builds from funnels via a GA4 internal-traffic filter."),
    ("User property: consent_version", "FNV-1a hash of the canonical terms text; bumps force re-consent."),
    ("User property: locale", "en | … (India/Brazil, English-only at MVP)."),
    ("User property: compat_passed", "true|false — segment funnels by device capability."),
    ("Auto param: network_type", "wifi | cellular | offline on every event (wrapper-injected)."),
    ("Auto param: is_practice", "true|false on every recording/upload event — practice captures are refused at enqueue (D-08) and must be separable."),
    ("PII — ALLOWED", "user_id, installation_id, recording_id, segment_id, task_id, durations, byte sizes, counts, numeric statuses, network_type, reasons/slugs."),
    ("PII — FORBIDDEN", "name, email, task NAME, search QUERY text, recording filename, lat/lng coordinates. (Existing T-2.4-01 guard — keep it.)"),
    ("GA4 limits", "Event name <=40 chars; <=25 params/event; <=500 distinct event names; param value <=100 chars. All names above comply."),
    ("Naming", "snake_case. Existing 47 allowlist names kept verbatim (additive only) — see Open Questions Q5 if you'd rather rename for consistency before GA4 ships."),
]
r = 4
for k, v in conv:
    kc = ws.cell(row=r, column=1, value=k); kc.font = MONO_B; kc.alignment = TOPL
    kc.fill = PatternFill("solid", fgColor="EEF1F6"); kc.border = BORDER
    vc = ws.cell(row=r, column=2, value=v); vc.font = T_FONT; vc.alignment = TOPL; vc.border = BORDER
    ws.row_dimensions[r].height = 32
    r += 1
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 110

# ===========================================================================
# SHEET — Open Questions
# ===========================================================================
ws = wb.create_sheet("Open Questions")
ws.sheet_view.showGridLines = False
QH = ["#", "Question", "My default / recommendation", "Why it matters", "Your decision"]
QW = [4, 46, 46, 46, 26]
q = [
    ("1", "Consent timing: GA4 collects from first_open, BEFORE the consent modal. Default analytics OFF until consent_agreed?",
         "Collection OFF until consent (analytics_collection_enabled=false), accept losing splash_shown -> consent_shown in GA4 (still in local telemetry ring).",
         "India DPDP / Brazil LGPD. Highest-priority decision — affects what's even collectible.", ""),
    ("2", "Should capture-QA telemetry (drift_max_ms, drift_p99_ms, mean_fps) ride on segment_finalized?",
         "Yes — gives the fleet-health drift dashboard the 2026-05-12 owner decision wanted with no extra pipeline.",
         "Otherwise events stay purely behavioral and drift lives only in metadata.json.", ""),
    ("3", "Practice scoping: same event names with is_practice=true, separate practice_* namespace, or excluded?",
         "Same names + is_practice flag — lets you compare practice vs real gate-pass rates.",
         "Affects every recording/upload event's cardinality.", ""),
    ("4", "Server-event transport: Postgres analytics_events table (BigQuery join) or GA4 Measurement Protocol?",
         "Postgres table joined in BigQuery.",
         "MP lacks geo/device context; decides whether server-step funnels are built in GA4 UI or SQL.", ""),
    ("5", "Allowlist hygiene: keep all 47 names verbatim (additive) or rename for consistency before GA4 ships?",
         "Keep verbatim, additive only — no churn.",
         "Renaming is cheap now, expensive once data accumulates in GA4.", ""),
    ("6", "Exclude dev/internal traffic (debug builds + team devices) from GA4?",
         "Yes — app_flavor user property + GA4 internal-traffic filter.",
         "With a small user base, ~5 test phones distort every funnel.", ""),
    ("7", "Capture 'attempt' definition: one record_start_pressed, or one Recording-screen visit?",
         "Recording-screen visit (recording_screen_opened) as the funnel entry, record_start_pressed as a step.",
         "Someone who rotates back to portrait and leaves never presses Start — invisible if attempts start at Start.", ""),
    ("8", "Eviction attribution: flag the EVICTING sign-in too (evicted_previous=true)?",
         "Yes — included on srv_user_signed_in.",
         "Distinguishes 'user upgraded phones' from credential sharing.", ""),
]
r = make_table(ws, QH, q, QW, start_row=1, mono_col=-1)
# taller rows for questions
for rr in range(2, r):
    ws.row_dimensions[rr].height = 64

wb.save(OUT)
print("WROTE", OUT)
print("Sheets:", wb.sheetnames)
